import pandas as pd
import openpyxl

def set_headers(ws, headers):
    for i in range(len(headers)):
        ws.cell(row=1, column=i+1, value=headers[i])

filename = "П-2 (инвест)_21апреля.xlsx"
code_column = "Код показателя"
OKPO_column = "ОКПО объекта"
sources_column = "Источники инвестиций  "
direction_column = "Направления инвестиций "
OKOF_column = "ОКОФ "
value_column = "Значение"
OKATO_column = "Код ОКАТО фактический"

codes = ['1141100020001_384_R1_1', '1141100020001_384_R3_1', '1141100020001_384_R3_2', '1141100020001_384_R2']
codes_and_sources_groups = {
    1: [['1141100020001_384_R1_1', 8], ['1141100020001_384_R3_1', 8], ['1141100020001_384_R3_2', 'any']],
    2: [['1141100020001_384_R2', 4], ['1141100020001_384_R3_1', 4]],
    3: [['1141100020001_384_R2', 11], ['1141100020001_384_R3_1', 11]],
    4: [['1141100020001_384_R2', 12], ['1141100020001_384_R3_1', 12]],
    5: [['1141100020001_384_R2', 13], ['1141100020001_384_R3_1', 13]],
    6: [['1141100020001_384_R2', 1], ['1141100020001_384_R3_1', 1]]
}

# чтение файла с заменой NaN на пустую строку
df = pd.read_excel(filename).fillna("")

# берем только колонки и преобразуем в массив
# в массив записываются все строки последовательно, в том же порядке, что и в файле
# номер строки в файле = индекс в массиве + 2
rows = df[[OKPO_column, code_column, sources_column, direction_column, OKOF_column, value_column, OKATO_column]].values.tolist()

result = {}
for index in range(len(rows)):
    [OKPO, code, source, direction, OKOF, value, OKATO] = rows[index]

    if code in codes:
        for group, codes_and_sources in codes_and_sources_groups.items():
            if ([code, source] in codes_and_sources or [code, 'any'] in codes_and_sources) \
                    and (code != '1141100020001_384_R1_1' or (direction == 18 and OKOF == '100.00.01.АГ')):
                if OKPO not in result:
                    result[OKPO] = {1: [], 2: [], 3: [], 4: [], 5: [], 6: []}

                result[OKPO][group].append([code, source, value, index + 2, OKATO])

# for key in result:
#     print(key)
#     print(result[key])
#     break

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'Весь результат'
set_headers(ws1, [OKPO_column, code_column, sources_column, value_column, 'номер строки в исходной таблице', OKATO_column])

# заполняем таблицу данными из словаря
row = 2
for OKPO, OKPO_data in result.items():
    for group, group_data in OKPO_data.items():
        if group_data:
            row += 1
            ws1.cell(row=row, column=1, value=OKPO)
        for data in group_data:
            ws1.cell(row=row, column=2, value=data[0])
            ws1.cell(row=row, column=3, value=data[1])
            ws1.cell(row=row, column=4, value=data[2])
            ws1.cell(row=row, column=5, value=data[3])
            ws1.cell(row=row, column=6, value=data[4])
            row += 1


ws2 = wb.create_sheet(title='Сокращенный результат')
set_headers(ws2, [OKPO_column, code_column, sources_column, value_column, 'номер строки в исходной таблице', 'В чем ошибка', OKATO_column])

row = 2
for OKPO, OKPO_data in result.items():
    for group, group_data in OKPO_data.items():
        correct_group = True
        correct_values = True
        error_message = ''
        value = 'qwe'
        for data in group_data:
            if value == 'qwe':
                value = data[2]
            elif value != data[2]:
                correct_values = False

        if group == 1 and len(group_data) == 3 or group != 1 and len(group_data) == 2:
            if not correct_values:
                correct_group = False
                error_message = 'Значения не совпадают'
        else:
            if correct_values:
                correct_group = False
                error_message = 'Неверное количество кодов показателя'
            else:
                correct_group = False
                error_message = 'Значения не совпадают и неверное количество кодов показателя'

        if not correct_group:
            if group_data:
                row += 1
                ws2.cell(row=row, column=1, value=OKPO)
                ws2.cell(row=row, column=6, value=error_message)
            for data in group_data:
                ws2.cell(row=row, column=2, value=data[0])
                ws2.cell(row=row, column=3, value=data[1])
                ws2.cell(row=row, column=4, value=data[2])
                ws2.cell(row=row, column=5, value=data[3])
                ws2.cell(row=row, column=7, value=data[4])
                row += 1

# сохраняем файл
wb.save('result.xlsx')
